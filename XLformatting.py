# Mary Catherine Shepherd, Sam Jenson, Gavin Clifton, Ben Funk, Thomas Apke
# professor anderson, section 003


import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

#insert path to the file!!
myWorkbook = openpyxl.load_workbook("/Users/samjenson/Desktop/poorlyorganized.xlsx")
# setting up the worksheet to be the active sheet in our file
currentSheet = myWorkbook.active


# setting up variables
colNum = 1

rowNum = 2
# creating a for loop to split up names and ID, and to create new sheets
for row in currentSheet.iter_rows(min_row=rowNum, values_only=True):
    classTitle = row[0]
    lstStudInfo = row[1].split('_')
    lstStudInfo.append(row[2])

    if classTitle not in myWorkbook.sheetnames:
        worksheet = myWorkbook.create_sheet(title=classTitle)
        printRow = 1
        myWorkbook[classTitle]['A1'] = 'First Name'
        myWorkbook[classTitle]['B1'] = 'Last Name'
        myWorkbook[classTitle]['C1'] = 'Student ID'
        myWorkbook[classTitle]['D1'] = 'Grade'

    #move down a row per iteration
    printRow += 1
    #print values on the new sheet as long as the classTitle matches
    # myWorkbook[classTitle]['A' + str(printRow)].append(lstStudInfo)
    myWorkbook[classTitle]['A' + str(printRow)] = lstStudInfo[0]
    myWorkbook[classTitle]['B' + str(printRow)] = lstStudInfo[1]
    myWorkbook[classTitle]['C' + str(printRow)] = lstStudInfo[2]
    myWorkbook[classTitle]['D' + str(printRow)] = lstStudInfo[3]
    myWorkbook[classTitle]['F1'] = 'Summary Statistics'
    myWorkbook[classTitle]['F2'] = 'Highest Grade'
    myWorkbook[classTitle]['F3'] = 'Lowest Grade'
    myWorkbook[classTitle]['F4'] = 'Mean Grade'
    myWorkbook[classTitle]['F5'] = 'Median Grade'
    myWorkbook[classTitle]['F6'] = 'Number of Students'
    myWorkbook[classTitle]['G1'] = 'Value'
    myWorkbook[classTitle]['G2'] = '=MAX(D2:D100)'
    myWorkbook[classTitle]['G3'] = '=MIN(D2:D100)'
    myWorkbook[classTitle]['G4'] = '=AVERAGE(D2:D100)'
    myWorkbook[classTitle]['G5'] = '=MEDIAN(D2:D100)'
    myWorkbook[classTitle]['G6'] = '=COUNT(D2:D100)'




#apply filters for each worksheet
for worksheet in myWorkbook.worksheets:
    max_row = worksheet.max_row
    worksheet.auto_filter.ref = f"A1:D{max_row}"

# formats each column with bold font and proper size
# Corrected formatting section - applies to ALL worksheets
for worksheet in myWorkbook.worksheets:
    # Format headers
    for col in ['A', 'B', 'C', 'D', 'F', 'G']:
        header_cell = worksheet[f'{col}1']
        header_cell.font = Font(bold=True)
        # Get length of header text or 10 as default
        header_length = len(str(header_cell.value)) if header_cell.value else 10
        worksheet.column_dimensions[col].width = header_length + 5


# Saves the new file
myWorkbook.save('formatted_grades.xlsx')
